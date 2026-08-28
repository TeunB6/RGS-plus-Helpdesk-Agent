"""import_check -- tell a user why their RGS+ Excel import did not do what they expected.

  import_list_templates      which templates are known, and their fields
  import_describe_template   the documented rules for one template
  import_validate_file       check a filled-in workbook against its own spec

WHY THIS IS A TOOL AND NOT A PROMPT
-----------------------------------
The RGS+ importer FAILS SILENTLY, by documented design. From objects.xlsx's own
uitleg sheet:

  * "Bij ontbrekende waarde of een tekst bij vhe wordt de regel OVERGESLAGEN met import."
  * "De kolom [type] is verplicht. Bij ontbrekende of foutieve waarde wordt de
     invoer 'UTILITEIT'."
  * "Bij een ontbrekende waarde of foutief geschreven naam wordt het veld NIET GEWIJZIGD."

No error is ever raised. Rows vanish, types change themselves, links never
happen -- and the user has no route back to the cause. That is why they write
"ik ben zelf aan het puzzelen en ik krijg mijn import niet voor elkaar" and
cannot say more.

Finding those faults is arithmetic over a spec, not judgement, so a model should
not be doing it: it would have to eyeball hundreds of cells and would miss some.
The tool finds them deterministically; the agent's job is to explain the result
in plain Dutch. Brian's stated doubt was "ik heb geen idee of een AI-agent dat
soort dingen eruit kan filteren" -- the honest answer is that it does not have
to.

THE RULES ARE NOT HAND-WRITTEN
------------------------------
Every RGS+ template ships an `uitleg` sheet holding a literal field table, so
the template IS the spec and this reads it out. Point it at a new template and
that one validates too, with no code change.

Two layouts exist and they are NOT interchangeable:

  A   Header | Type | Verplicht | Uitleg      adressen, objects, pricebook,
                                              inspectionlist
  B   header | import | omschrijving          scenario

Layout B has no Type column, so type checking is simply unavailable there --
better than inferring a type from a column name and inventing an error. It also
uses "n.v.t." for fields that appear in an EXPORT but are ignored on import,
which is not the same thing as optional.

TWO BEHAVIOURS THAT LOOK LIKE BUGS AND ARE NOT
----------------------------------------------
1. It REFUSES to run when it cannot parse a spec, rather than reporting the file
   as clean. A validator with no rules finds no problems, which is the most
   dangerous output it could give.

2. Quoted values are only a CLOSED set when the wording says so. scenario's
   `laag` reads: 'laag van element OF "inspectie" / "staartkosten"' -- the layer
   name is free text and those two are merely special values for indirect costs.
   Reading that as an enum flagged every correct row in the file. A validator
   that cries wolf gets switched off, which is worse than missing a rule.

Requires: openpyxl (pip install openpyxl).
Optional env: RGSPLUS_UPLOAD_DIR -- directory holding user-supplied workbooks
              (default ~/.hermes/uploads). Files outside it are refused.
"""

from __future__ import annotations

import os
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _OPENPYXL_ERROR = None
except ImportError as e:                                  # pragma: no cover
    load_workbook = None                                   # type: ignore[assignment]
    get_column_letter = None                               # type: ignore[assignment]
    _OPENPYXL_ERROR = str(e)

REQUIRED_WORDS = {"ja", "verplicht", "verplicht*"}

TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"

#: A systematic fault (a whole column pasted as text) produces one problem per
#: row. Reporting 5.000 of them to a language model is useless and expensive, so
#: identical problems are collapsed and the list is capped. The COUNT is always
#: reported honestly -- silent truncation would read as "that's all of them".
MAX_REPORTED = 40
MAX_PER_KIND = 5


# ---------------------------------------------------------------------------
# Spec parsing -- read the field table out of the uitleg sheet(s)
# ---------------------------------------------------------------------------

class Field:
    def __init__(self, name, type_, verplicht, uitleg):
        self.name = (name or "").strip()
        self.key = self.name.lower()
        self.type = (type_ or "").strip().lower()
        self.required = (verplicht or "").strip().lower() in REQUIRED_WORDS
        self.uitleg = (uitleg or "").strip()
        self.enum = self._enum()
        self.minimum = self._minimum()
        self.range = self._range()
        self.case_sensitive = bool(
            re.search(r"hoofdletter\s*gevoelig|exact overeen", self.uitleg, re.I)
        )
        self.ignored = False  # layout B "n.v.t." -- in the export, ignored on import

    def _enum(self):
        # 'N (nul) - H (hoog) - L (laag)' -- scenario.xlsx spells its options out
        # longhand, and omits V because "een import met BTW verdelen is niet
        # mogelijk". The pricebook's four options are only three here. Read each
        # template's own words; never share an enum between templates.
        letters = re.findall(r"\b([A-Z])\s*\((?:nul|hoog|laag|verdelen)\)", self.uitleg)
        if len(letters) >= 2:
            return letters
        # Closed only when the wording says so -- see module docstring, note 2.
        quoted = re.findall(r'"([^"]+)"', self.uitleg)
        closed = re.search(r"\b(waarde|keuze uit)\b", self.uitleg, re.I) or \
            self.uitleg.lstrip().startswith('"')
        if len(quoted) >= 2 and closed:
            return [q for q in quoted if len(q) < 25]
        m = re.search(r"keuze uit\s+([A-Za-z0-9/\s]+)", self.uitleg, re.I)
        if m:
            parts = [p.strip() for p in m.group(1).split("/") if p.strip()]
            if len(parts) >= 2:
                return parts
        return None

    def _minimum(self):
        m = re.search(r"minimaal\s+(\d+)", self.uitleg, re.I)
        return int(m.group(1)) if m else None

    def _range(self):
        m = re.search(r"tussen de\s+(\d+)\s+en\s+(\d+)", self.uitleg, re.I)
        return (int(m.group(1)), int(m.group(2))) if m else None

    @property
    def numeric(self):
        return self.type.startswith("num")


def _read_spec(path):
    """Return {key: Field} from every `uitleg*` sheet in the workbook."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        fields = {}
        for name in wb.sheetnames:
            if not name.lower().startswith("uitleg"):
                continue
            layout = None
            for row in wb[name].iter_rows(values_only=True):
                cells = [("" if c is None else str(c).strip()) for c in row]
                head = [c.lower() for c in cells[:4]]
                if head[:4] == ["header", "type", "verplicht", "uitleg"]:
                    layout = "A"
                    continue
                if head[:3] == ["header", "import", "omschrijving"]:
                    layout = "B"
                    continue
                if layout == "A":
                    if not cells[0] or cells[0].startswith("…") or not cells[1]:
                        layout = None
                        continue
                    f = Field(*(cells + ["", "", "", ""])[:4])
                    fields.setdefault(f.key, f)
                elif layout == "B":
                    if not cells[0] or cells[0].startswith("…"):
                        layout = None
                        continue
                    imp = (cells[1] if len(cells) > 1 else "").lower()
                    f = Field(
                        cells[0], "",
                        "verplicht" if imp == "verplicht" else "nee",
                        cells[2] if len(cells) > 2 else "",
                    )
                    f.ignored = imp.startswith("n.v.t")
                    fields.setdefault(f.key, f)
        return fields
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Path safety
# ---------------------------------------------------------------------------

def _upload_dir() -> Path:
    raw = os.environ.get("RGSPLUS_UPLOAD_DIR", "").strip()
    return Path(raw).expanduser() if raw else Path.home() / ".hermes" / "uploads"


def _resolve(path_str: str) -> tuple[Path | None, str | None]:
    """Resolve a caller-supplied path, refusing anything outside the allowed roots.

    The file being checked is uploaded by a customer, and the path arrives via a
    model that just read that customer's text. Without this, "validate
    ../../.env" is a question a user could ask.
    """
    if not path_str or not path_str.strip():
        return None, "No file given."
    try:
        candidate = Path(path_str).expanduser().resolve()
    except (OSError, RuntimeError) as e:
        return None, f"Could not resolve that path: {e}"

    roots = [_upload_dir().resolve()]
    if TEMPLATE_DIR.exists():
        roots.append(TEMPLATE_DIR.resolve())

    for root in roots:
        try:
            candidate.relative_to(root)
            break
        except ValueError:
            continue
    else:
        allowed = ", ".join(str(r) for r in roots)
        return None, (
            f"Refusing to read {candidate}: outside the allowed directories "
            f"({allowed}). Uploaded workbooks belong in the upload directory."
        )

    if not candidate.is_file():
        return None, f"No such file: {candidate.name}"
    if candidate.suffix.lower() not in (".xlsx", ".xlsm"):
        return None, f"{candidate.name} is not an .xlsx file."
    return candidate, None


def _known_templates() -> dict[str, Path]:
    if not TEMPLATE_DIR.exists():
        return {}
    return {p.stem.lower(): p for p in sorted(TEMPLATE_DIR.glob("*.xlsx"))}


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate(spec_fields, target_path):
    """Return a list of (cell_ref | None, field_name, what, consequence)."""
    wb = load_workbook(target_path, data_only=True, read_only=True)
    try:
        first = wb.sheetnames[0]

        # Every template says "zorg dat het import tabblad vooraan staat !!!"
        # because the importer reads the FIRST sheet and nothing else. A user who
        # reorders the tabs silently imports the instructions as data.
        if first.lower().startswith("uitleg"):
            return [(
                None, first,
                f"The first sheet is '{first}' -- the instructions, not your data.",
                "RGS+ imports the first sheet only, so it would read the "
                "instruction text as rows. Drag the data sheet to the front and save.",
            )]

        ws = wb[first]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [(None, first, "The first sheet is empty.", "Nothing would import.")]

        headers = [("" if c is None else str(c).strip()) for c in rows[0]]
        index = {h.lower(): i for i, h in enumerate(headers) if h}

        problems = []
        for f in spec_fields.values():
            if f.required and not f.ignored and f.key not in index:
                problems.append((
                    None, f.name,
                    f"Required column '{f.name}' is missing.",
                    f.uitleg or "Rows will be skipped or silently defaulted.",
                ))

        for r, row in enumerate(rows[1:], start=2):
            for key, col in index.items():
                f = spec_fields.get(key)
                if not f or f.ignored:
                    continue
                raw = row[col] if col < len(row) else None
                ref = f"{get_column_letter(col + 1)}{r}"

                if raw is None or str(raw).strip() == "":
                    if f.required:
                        problems.append((
                            ref, f.name, f"'{f.name}' is empty.",
                            f.uitleg or "This row will be skipped on import.",
                        ))
                    continue

                val = str(raw).strip()

                if f.numeric and isinstance(raw, str):
                    problems.append((
                        ref, f.name,
                        f"'{val}' is stored as TEXT but '{f.name}' must be numeric.",
                        "Excel keeps pasted values and leading zeros as text. "
                        "RGS+ will skip this row without reporting it.",
                    ))
                elif f.numeric:
                    try:
                        n = float(raw)
                        if f.minimum is not None and n < f.minimum:
                            problems.append((
                                ref, f.name,
                                f"{n:g} is below the minimum of {f.minimum}.", f.uitleg))
                        if f.range and not (f.range[0] <= n <= f.range[1]):
                            problems.append((
                                ref, f.name,
                                f"{n:g} is outside {f.range[0]}-{f.range[1]}.", f.uitleg))
                    except (TypeError, ValueError):
                        pass

                if f.enum:
                    hit = (val in f.enum if f.case_sensitive
                           else val.lower() in [e.lower() for e in f.enum])
                    if not hit:
                        near = [e for e in f.enum if e.lower() == val.lower()]
                        if near:
                            problems.append((
                                ref, f.name,
                                f"'{val}' should be '{near[0]}' -- this field is "
                                f"case-sensitive.", f.uitleg))
                        else:
                            problems.append((
                                ref, f.name,
                                f"'{val}' is not one of: {', '.join(f.enum)}.", f.uitleg))
        return problems
    finally:
        wb.close()


def _render(problems, spec_name, file_name, field_count, spec_source):
    lines = [
        f"file: {file_name}",
        f"spec: {spec_name} ({field_count} documented fields, read from {spec_source})",
    ]
    if not problems:
        lines.append("")
        lines.append("OK -- nothing found. This file should import as expected.")
        return "\n".join(lines)

    # Collapse repeats: one systematic fault should read as one fault.
    grouped: dict[tuple[str, str], list] = {}
    for ref, field, what, why in problems:
        kind = re.sub(r"'[^']*'", "'…'", what)
        grouped.setdefault((field, kind), []).append((ref, what, why))

    lines.append("")
    lines.append(f"{len(problems)} problem(s) in {len(grouped)} distinct kind(s):")
    lines.append("")

    shown = 0
    for (field, _kind), items in grouped.items():
        if shown >= MAX_REPORTED:
            lines.append(f"  … and {len(problems) - shown} more, not listed.")
            break
        head = items[:MAX_PER_KIND]
        for ref, what, why in head:
            where = f"cell {ref}" if ref else f"column '{field}'"
            lines.append(f"  - {where}: {what}")
            if why:
                lines.append(f"      why it matters: {why}")
            shown += 1
        if len(items) > len(head):
            extra = len(items) - len(head)
            refs = ", ".join(i[0] for i in items[len(head):len(head) + 8] if i[0])
            lines.append(f"      same problem in {extra} more row(s): {refs}"
                         + (" …" if extra > 8 else ""))
            shown += extra
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------

def _guard():
    if _OPENPYXL_ERROR:
        return f"ERROR: openpyxl is not installed in the agent image ({_OPENPYXL_ERROR})."
    return None


_LIST_SCHEMA = {
    "name": "import_list_templates",
    "description": (
        "List the RGS+ import templates this tool knows, with how many fields "
        "each documents. Call this when a customer mentions an import but you "
        "are not sure which template they mean."
    ),
    "parameters": {"type": "object", "properties": {}, "required": []},
}


def _handle_list(params, **kwargs):
    err = _guard()
    if err:
        return err
    templates = _known_templates()
    if not templates:
        return "No bundled templates found."
    out = ["Known RGS+ import templates:"]
    for name, path in templates.items():
        try:
            out.append(f"  {name} -- {len(_read_spec(path))} documented fields")
        except Exception as e:                              # noqa: BLE001
            out.append(f"  {name} -- could not be read ({e})")
    out.append("")
    out.append("The importer reads the FIRST sheet only; the uitleg sheet is the spec.")
    return "\n".join(out)


_DESCRIBE_SCHEMA = {
    "name": "import_describe_template",
    "description": (
        "Show the documented rules for one RGS+ import template -- which columns "
        "are mandatory, their types, allowed values and case-sensitivity. Use it "
        "to answer 'welke velden zijn verplicht bij X' from the template itself "
        "rather than from memory."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "template": {
                "type": "string",
                "description": "Template name, e.g. objects, adressen, pricebook, "
                               "inspectionlist, scenario, structure.",
            },
        },
        "required": ["template"],
    },
}


def _handle_describe(params, **kwargs):
    err = _guard()
    if err:
        return err
    name = str(params.get("template", "")).strip().lower()
    templates = _known_templates()
    if name not in templates:
        return (f"Unknown template {name!r}. Known: "
                f"{', '.join(sorted(templates)) or 'none'}.")
    fields = _read_spec(templates[name])
    if not fields:
        return f"No field table found in {name} -- its uitleg sheet uses an unknown layout."

    out = [f"{name} -- {len(fields)} documented fields", ""]
    for f in fields.values():
        bits = []
        if f.required:
            bits.append("REQUIRED")
        if f.ignored:
            bits.append("ignored on import (export only)")
        if f.type:
            bits.append(f.type)
        if f.enum:
            bits.append("one of: " + ", ".join(f.enum))
        if f.minimum is not None:
            bits.append(f"min {f.minimum}")
        if f.range:
            bits.append(f"range {f.range[0]}-{f.range[1]}")
        if f.case_sensitive:
            bits.append("case-sensitive")
        out.append(f"  {f.name}" + (f"  [{'; '.join(bits)}]" if bits else ""))
        if f.uitleg:
            out.append(f"      {f.uitleg}")
    return "\n".join(out)


_VALIDATE_SCHEMA = {
    "name": "import_validate_file",
    "description": (
        "Check a filled-in RGS+ import workbook and report exactly what would go "
        "wrong, and what the CONSEQUENCE would be. Use this whenever a customer "
        "says their import did not work, or asks you to check a file before they "
        "import it -- the RGS+ importer fails silently, so 'nothing happened' is "
        "the normal symptom and this is the only way to find the cause. "
        "Report the findings in plain Dutch; do not paste the raw output."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Path to the customer's .xlsx, inside the upload directory.",
            },
            "template": {
                "type": "string",
                "description": (
                    "Optional. Only needed when the customer's workbook no longer "
                    "contains its own uitleg sheet -- normally it does, and the "
                    "file is its own spec."
                ),
            },
        },
        "required": ["file_path"],
    },
}


def _handle_validate(params, **kwargs):
    err = _guard()
    if err:
        return err

    target, err = _resolve(str(params.get("file_path", "")))
    if err:
        return f"ERROR: {err}"

    # The customer's own workbook is usually the template they downloaded, with
    # the uitleg sheet still in it -- so it carries its own spec and stays
    # correct even if RGS+ revises the template. Bundled copies are the fallback.
    spec_source, spec_name = "the file's own uitleg sheet", target.name
    try:
        fields = _read_spec(target)
    except Exception as e:                                  # noqa: BLE001
        return f"ERROR: could not read {target.name}: {e}"

    if not fields:
        wanted = str(params.get("template", "")).strip().lower()
        templates = _known_templates()
        if not wanted:
            return (
                f"{target.name} has no `uitleg` sheet, so it does not carry its own "
                f"rules, and no template was named. Ask the customer which import "
                f"this is and call again with `template`. Known: "
                f"{', '.join(sorted(templates)) or 'none'}."
            )
        if wanted not in templates:
            return (f"Unknown template {wanted!r}. Known: "
                    f"{', '.join(sorted(templates)) or 'none'}.")
        try:
            fields = _read_spec(templates[wanted])
        except Exception as e:                              # noqa: BLE001
            return f"ERROR: could not read template {wanted}: {e}"
        spec_source, spec_name = "the bundled template", wanted

    # Refuse rather than bless. Zero fields means zero rules, and a validator
    # with no rules reports every file as clean -- the worst possible answer.
    if not fields:
        return (
            f"ERROR: no field table could be read for {target.name}, so there are no "
            f"rules to check against. Refusing to report this file as clean -- that "
            f"would be a false all-clear. The uitleg sheet uses a layout this tool "
            f"does not know; report it rather than guessing."
        )

    try:
        problems = _validate(fields, target)
    except Exception as e:                                  # noqa: BLE001
        return f"ERROR: could not validate {target.name}: {e}"

    return _render(problems, spec_name, target.name, len(fields), spec_source)


def register(ctx):
    for name, schema, handler in (
        ("import_list_templates", _LIST_SCHEMA, _handle_list),
        ("import_describe_template", _DESCRIBE_SCHEMA, _handle_describe),
        ("import_validate_file", _VALIDATE_SCHEMA, _handle_validate),
    ):
        ctx.register_tool(name=name, toolset="import_check", schema=schema, handler=handler)
